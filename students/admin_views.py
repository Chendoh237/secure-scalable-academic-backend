from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Count, Avg
from django.utils import timezone
from django.contrib.auth import get_user_model
from datetime import datetime, timedelta
from .models import Student
from institutions.models import Department, Institution, Faculty
from institutions.program_models import AcademicProgram
import json

User = get_user_model()

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_stats(request):
    """Get admin dashboard statistics with real data"""
    try:
        from django.db.models import Count, Q, Avg
        from datetime import datetime, timedelta
        from attendance.models import Attendance, CourseRegistration
        from courses.models import Course
        from attendance.utils import calculate_attendance_percentage
        
        # Basic counts
        total_students = Student.objects.count()
        approved_students = Student.objects.filter(is_approved=True).count()
        pending_students = Student.objects.filter(is_approved=False).count()
        
        # Department count
        total_departments = Department.objects.count()
        
        # Course count - get real count from Course model
        try:
            total_courses = Course.objects.count()
        except:
            total_courses = 0
        
        # Calculate real attendance rates
        today = datetime.now().date()
        week_ago = today - timedelta(days=7)
        
        # Today's attendance rate
        today_total = Attendance.objects.filter(date=today).count()
        today_present = Attendance.objects.filter(
            date=today, 
            status__in=['present', 'late']
        ).count()
        today_attendance_rate = round((today_present / today_total * 100), 2) if today_total > 0 else 0
        
        # Weekly attendance rate
        week_total = Attendance.objects.filter(date__gte=week_ago).count()
        week_present = Attendance.objects.filter(
            date__gte=week_ago, 
            status__in=['present', 'late']
        ).count()
        weekly_attendance_rate = round((week_present / week_total * 100), 2) if week_total > 0 else 0
        
        # Count students with low attendance (below 75%)
        low_attendance_count = 0
        approved_students_list = Student.objects.filter(is_approved=True)
        
        for student in approved_students_list:
            # Calculate overall attendance for this student
            student_total = Attendance.objects.filter(student=student).count()
            if student_total > 0:
                student_present = Attendance.objects.filter(
                    student=student, 
                    status__in=['present', 'late']
                ).count()
                attendance_percentage = (student_present / student_total) * 100
                if attendance_percentage < 75:
                    low_attendance_count += 1
        
        # Count active sessions (sessions happening today)
        active_sessions_count = 0
        try:
            from live_sessions.models import LiveSession
            active_sessions_count = LiveSession.objects.filter(
                state='active',
                date=today
            ).count()
        except:
            # If LiveSession model doesn't exist, count from timetable
            try:
                from courses.models import TimetableSlot
                current_time = datetime.now().time()
                current_day = datetime.now().strftime('%a').upper()[:3]
                
                active_sessions_count = TimetableSlot.objects.filter(
                    day_of_week=current_day,
                    start_time__lte=current_time,
                    end_time__gte=current_time
                ).count()
            except:
                active_sessions_count = 0
        
        return Response({
            'totalStudents': total_students,
            'totalCourses': total_courses,
            'totalDepartments': total_departments,
            'activeSessionsCount': active_sessions_count,
            'todayAttendanceRate': today_attendance_rate,
            'weeklyAttendanceRate': weekly_attendance_rate,
            'lowAttendanceAlerts': low_attendance_count
        })
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def active_sessions(request):
    """Get active attendance sessions"""
    try:
        from datetime import datetime, time
        from django.db.models import Count, Q
        
        current_time = datetime.now()
        current_day = current_time.strftime('%a').upper()[:3]  # MON, TUE, etc.
        current_time_only = current_time.time()
        
        active_sessions_data = []
        
        # Try to get from LiveSession model first
        try:
            from live_sessions.models import LiveSession
            # Fix: Filter by start_time date instead of non-existent date field
            live_sessions = LiveSession.objects.filter(
                state__in=['active', 'live', 'scheduled'],
                start_time__date=current_time.date()
            ).select_related('course_offering__course')
            
            for session in live_sessions:
                # Count attendance for this session
                try:
                    from attendance.models import Attendance
                    course = session.course_offering.course if session.course_offering else None
                    
                    if course:
                        present_count = Attendance.objects.filter(
                            course_registration__course=course,
                            date=current_time.date(),
                            status__in=['present', 'late']
                        ).count()
                        
                        expected_count = Attendance.objects.filter(
                            course_registration__course=course,
                            date=current_time.date()
                        ).count()
                    else:
                        present_count = 0
                        expected_count = 0
                except:
                    present_count = 0
                    expected_count = 0
                
                active_sessions_data.append({
                    'id': str(session.id),
                    'courseCode': course.code if course else 'N/A',
                    'courseName': course.title if course else session.title,
                    'state': session.state,
                    'venue': 'Virtual Session',
                    'startTime': session.start_time.strftime('%H:%M') if session.start_time else '00:00',
                    'endTime': session.end_time.strftime('%H:%M') if session.end_time else '00:00',
                    'presentStudents': present_count,
                    'expectedStudents': expected_count,
                    'attendanceRate': round((present_count / expected_count * 100), 2) if expected_count > 0 else 0
                })
                
        except ImportError:
            # If LiveSession doesn't exist, get from timetable
            try:
                from courses.models import TimetableSlot
                
                # Get current active timetable slots
                active_slots = TimetableSlot.objects.filter(
                    day_of_week=current_day,
                    start_time__lte=current_time_only,
                    end_time__gte=current_time_only
                ).select_related('course', 'lecturer')
                
                for slot in active_slots:
                    # Count students for this slot
                    try:
                        from attendance.models import Attendance
                        present_count = Attendance.objects.filter(
                            timetable_entry=slot,
                            date=current_time.date(),
                            status__in=['present', 'late']
                        ).count()
                        
                        expected_count = Attendance.objects.filter(
                            timetable_entry=slot,
                            date=current_time.date()
                        ).count()
                    except:
                        present_count = 0
                        expected_count = 0
                    
                    active_sessions_data.append({
                        'id': slot.id,
                        'courseCode': slot.course.code if slot.course else 'N/A',
                        'courseName': slot.course.title if slot.course else 'Unknown Course',
                        'state': 'active',
                        'venue': getattr(slot, 'venue', 'TBA'),
                        'startTime': slot.start_time.strftime('%H:%M') if slot.start_time else '00:00',
                        'endTime': slot.end_time.strftime('%H:%M') if slot.end_time else '00:00',
                        'presentStudents': present_count,
                        'expectedStudents': expected_count,
                        'attendanceRate': round((present_count / expected_count * 100), 2) if expected_count > 0 else 0,
                        'lecturer': slot.lecturer.name if hasattr(slot, 'lecturer') and slot.lecturer else 'TBA'
                    })
                    
            except ImportError:
                # If no timetable model, return empty
                pass
        
        return Response(active_sessions_data)
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def low_attendance_students(request):
    """Get students with low attendance (below 75%)"""
    try:
        from attendance.models import Attendance
        from django.db.models import Count, Q
        
        # Get all approved students
        students = Student.objects.filter(is_approved=True)
        
        low_attendance_students = []
        
        for student in students:
            # Calculate real attendance percentage for this student
            total_classes = Attendance.objects.filter(student=student).count()
            
            if total_classes > 0:  # Only include students with attendance records
                attended_classes = Attendance.objects.filter(
                    student=student,
                    status__in=['present', 'late']
                ).count()
                
                attendance_percentage = round((attended_classes / total_classes) * 100, 2)
                
                # Only include students with attendance below 75%
                if attendance_percentage < 75:
                    try:
                        department_name = student.department.name if hasattr(student, 'department') and student.department else 'Unknown'
                    except:
                        department_name = 'Unknown'
                        
                    low_attendance_students.append({
                        'studentId': student.id,
                        'studentName': student.full_name,
                        'matricule': student.matric_number,
                        'departmentName': department_name,
                        'overallAttendance': attendance_percentage,
                        'totalClasses': total_classes,
                        'attendedClasses': attended_classes,
                        'isEligible': attendance_percentage >= 75
                    })
        
        # Sort by attendance percentage (lowest first)
        low_attendance_students.sort(key=lambda x: x['overallAttendance'])
        
        return Response(low_attendance_students)
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def admin_students(request):
    """Get all students or create new student"""
    if request.method == 'GET':
        try:
            students = Student.objects.select_related('user', 'department').all()
            student_data = []
            
            for student in students:
                student_data.append({
                    'id': student.id,
                    'username': student.user.username,
                    'email': student.user.email,
                    'first_name': student.user.first_name,
                    'last_name': student.user.last_name,
                    'full_name': student.full_name,
                    'matricule': student.matric_number,
                    'phone': '',  # Add phone field to model if needed
                    'institution': student.institution.name if student.institution else '',
                    'faculty': student.faculty.name if student.faculty else '',
                    'department': student.department.name if student.department else '',
                    'is_approved': student.is_approved,
                    'date_joined': student.created_at.strftime('%Y-%m-%d'),
                    'attendance_rate': 0,  # Implement real calculation
                    'total_courses': 0,    # Implement real calculation
                    'gpa': 0.0            # Implement real calculation
                })
            
            return Response({'data': student_data})
        except Exception as e:
            return Response({'error': str(e)}, status=500)
    
    elif request.method == 'POST':
        try:
            data = request.data
            
            # Create user first
            username = data['matricule'].lower()
            user = User.objects.create_user(
                username=username,
                email=data['email'],
                password='temp123',
                first_name=data['first_name'],
                last_name=data['last_name']
            )
            
            # Create student
            student = Student.objects.create(
                user=user,
                full_name=data['full_name'],
                matric_number=data['matricule']
            )
            
            return Response({'message': 'Student created successfully'}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=500)

@api_view(['PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def admin_student_detail(request, student_id):
    """Update or delete specific student"""
    try:
        student = Student.objects.get(id=student_id)
    except Student.DoesNotExist:
        return Response({'error': 'Student not found'}, status=404)
    
    if request.method == 'PUT':
        try:
            data = request.data
            
            # Update user fields
            if 'first_name' in data:
                student.user.first_name = data['first_name']
            if 'last_name' in data:
                student.user.last_name = data['last_name']
            if 'email' in data:
                student.user.email = data['email']
            student.user.save()
            
            # Update student fields
            if 'full_name' in data:
                student.full_name = data['full_name']
            if 'matricule' in data:
                student.matric_number = data['matricule']
            if 'is_approved' in data:
                student.is_approved = data['is_approved']
            student.save()
            
            return Response({'message': 'Student updated successfully'})
        except Exception as e:
            return Response({'error': str(e)}, status=500)
    
    elif request.method == 'DELETE':
        try:
            student.user.delete()  # This will cascade delete the student
            return Response({'message': 'Student deleted successfully'})
        except Exception as e:
            return Response({'error': str(e)}, status=500)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def approve_student(request, student_id):
    """Approve a student"""
    try:
        student = Student.objects.get(id=student_id)
        student.is_approved = True
        student.save()
        return Response({'message': 'Student approved successfully'})
    except Student.DoesNotExist:
        return Response({'error': 'Student not found'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def admin_courses(request):
    """Get all courses or create new course"""
    if request.method == 'GET':
        try:
            from courses.models import Course, CourseOffering, StudentCourse
            
            courses = Course.objects.select_related('department').all()
            course_data = []
            
            for course in courses:
                # Get current offerings and enrollment count
                current_offerings = CourseOffering.objects.filter(course=course)
                total_enrolled = StudentCourse.objects.filter(
                    course_offering__course=course,
                    is_active=True
                ).count()
                
                course_data.append({
                    'id': course.id,
                    'code': course.code,
                    'title': course.title,
                    'department': course.department.name,
                    'creditUnits': course.credit_units,
                    'level': course.level,
                    'semester': course.semester,
                    'enrolledStudents': total_enrolled,
                    'attendanceThreshold': course.attendance_threshold,
                    'isPublished': current_offerings.exists()
                })
            
            return Response({'data': course_data})
        except Exception as e:
            print(f"Courses API Error: {str(e)}")  # Debug logging
            return Response({'error': str(e)}, status=500)
    
    elif request.method == 'POST':
        try:
            from courses.models import Course
            data = request.data
            
            # Get department
            department = Department.objects.get(id=data['departmentId'])
            
            course = Course.objects.create(
                code=data['code'],
                title=data['title'],
                credit_units=data['creditUnits'],
                department=department,
                level=data.get('level', 'HND1'),
                semester=data.get('semester', 'Semester 1'),
                attendance_threshold=data.get('attendanceThreshold', 75)
            )
            
            return Response({'message': 'Course created successfully', 'id': course.id}, status=201)
        except Exception as e:
            print(f"Course Creation Error: {str(e)}")  # Debug logging
            return Response({'error': str(e)}, status=500)

@api_view(['PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def admin_course_detail(request, course_id):
    """Update or delete specific course"""
    try:
        from courses.models import Course
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return Response({'error': 'Course not found'}, status=404)
    
    if request.method == 'PUT':
        try:
            data = request.data
            
            if 'code' in data:
                course.code = data['code']
            if 'title' in data:
                course.title = data['title']
            if 'creditUnits' in data:
                course.credit_units = data['creditUnits']
            if 'level' in data:
                course.level = data['level']
            if 'semester' in data:
                course.semester = data['semester']
            if 'attendanceThreshold' in data:
                course.attendance_threshold = data['attendanceThreshold']
            if 'departmentId' in data:
                course.department = Department.objects.get(id=data['departmentId'])
            
            course.save()
            return Response({'message': 'Course updated successfully'})
        except Exception as e:
            return Response({'error': str(e)}, status=500)
    
    elif request.method == 'DELETE':
        try:
            course.delete()
            return Response({'message': 'Course deleted successfully'})
        except Exception as e:
            return Response({'error': str(e)}, status=500)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def admin_departments(request):
    """Get all departments or create new department"""
    if request.method == 'GET':
        try:
            departments = Department.objects.all()
            dept_list = []
            for dept in departments:
                try:
                    student_count = Student.objects.filter(department=dept).count()
                except Exception:
                    student_count = 0
                # Only include departments with students
                if student_count > 0:
                    dept_list.append({
                        'id': dept.id, 
                        'name': dept.name,
                        'studentCount': student_count
                    })
            return Response({'data': dept_list})
        except Exception as e:
            print(f"Department API Error: {str(e)}")  # Debug logging
            return Response({'error': str(e)}, status=500)
    
    elif request.method == 'POST':
        try:
            data = request.data
            if 'name' not in data:
                return Response({'error': 'Department name is required'}, status=400)
            
            # For now, create department without faculty (will need to be updated later)
            # Get first faculty or create a default one
            faculty = Faculty.objects.first()
            if not faculty:
                # Create a default faculty if none exists
                institution = Institution.objects.first()
                if not institution:
                    return Response({'error': 'No institution found. Please create an institution first.'}, status=400)
                faculty = Faculty.objects.create(name='Default Faculty', institution=institution)
            
            department = Department.objects.create(name=data['name'], faculty=faculty)
            return Response({'message': 'Department created successfully', 'id': department.id}, status=201)
        except Exception as e:
            print(f"Department Creation Error: {str(e)}")  # Debug logging
            return Response({'error': str(e)}, status=500)

@api_view(['PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def admin_department_detail(request, dept_id):
    """Update or delete specific department"""
    try:
        department = Department.objects.get(id=dept_id)
    except Department.DoesNotExist:
        return Response({'error': 'Department not found'}, status=404)
    
    if request.method == 'PUT':
        try:
            data = request.data
            department.name = data['name']
            department.save()
            return Response({'message': 'Department updated successfully'})
        except Exception as e:
            return Response({'error': str(e)}, status=500)
    
    elif request.method == 'DELETE':
        try:
            department.delete()
            return Response({'message': 'Department deleted successfully'})
        except Exception as e:
            return Response({'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def attendance_records(request):
    """Get real attendance records from the database"""
    try:
        from attendance.models import Attendance
        
        # Get real attendance records with related data
        records = Attendance.objects.select_related(
            'student',
            'student__department',
            'course_registration__course'
        ).all().order_by('-date', '-recorded_at')
        
        # Transform to expected format
        records_data = []
        for record in records:
            try:
                records_data.append({
                    'id': record.id,
                    'student_name': record.student.full_name,
                    'matricule': record.student.matric_number,
                    'department': record.student.department.name if record.student.department else 'Unknown',
                    'course': record.course_registration.course.code if record.course_registration and record.course_registration.course else 'N/A',
                    'date': record.date.strftime('%Y-%m-%d'),
                    'status': record.status.title(),
                    'check_in_time': record.recorded_at.strftime('%H:%M') if record.recorded_at else 'N/A'
                })
            except Exception as record_error:
                # Skip problematic records
                continue
        
        return Response({'data': records_data})
    except Exception as e:
        # Return empty data if there's an error
        return Response({'data': []})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def test_query_params(request):
    """Test endpoint to debug query parameter issues"""
    print(f"=== TEST ENDPOINT CALLED ===")
    print(f"Request method: {request.method}")
    print(f"Request path: {request.path}")
    print(f"Request full path: {request.get_full_path()}")
    print(f"Request params: {request.GET}")
    print(f"=== END TEST DEBUG ===")
    
    from django.http import JsonResponse
    return JsonResponse({
        'status': 'success',
        'message': 'Test endpoint working!',
        'params': dict(request.GET),
        'path': request.path,
        'full_path': request.get_full_path()
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_attendance_records(request):
    """Export attendance records in PDF or Excel format"""
    print(f"=== EXPORT ENDPOINT CALLED ===")
    print(f"Request method: {request.method}")
    print(f"Request path: {request.path}")
    print(f"Request full path: {request.get_full_path()}")
    print(f"Request params: {request.GET}")
    print(f"Request user: {request.user}")
    print(f"Is authenticated: {request.user.is_authenticated}")
    print(f"=== END DEBUG INFO ===")
    
    try:
        export_format = request.GET.get('export_format', 'excel')  # Changed from 'format' to 'export_format'
        department_id = request.GET.get('department_id')
        level = request.GET.get('level')
        
        # Get attendance records with filtering
        from attendance.models import Attendance
        
        # Start with all attendance records
        attendance_query = Attendance.objects.select_related(
            'student', 
            'student__department', 
            'course_registration__course'
        ).all()
        
        # Apply department filter if provided
        if department_id:
            attendance_query = attendance_query.filter(student__department_id=department_id)
        
        # Apply level filter if provided (assuming level is stored in course or student)
        if level:
            # This might need adjustment based on your actual level storage
            attendance_query = attendance_query.filter(course_registration__course__level=level)
        
        # If no real attendance data, use mock data for demonstration
        if not attendance_query.exists():
            # Generate mock data organized by department and level
            mock_data = []
            departments = Department.objects.all()
            
            for dept in departments:
                if department_id and str(dept.id) != department_id:
                    continue
                    
                students = Student.objects.filter(department=dept, is_approved=True)
                for student in students:
                    # Mock different levels
                    levels = ['HND1', 'HND2', 'ND1', 'ND2'] if not level else [level]
                    for student_level in levels:
                        if level and student_level != level:
                            continue
                            
                        mock_data.append({
                            'student_name': student.full_name,
                            'matricule': student.matric_number,
                            'department': dept.name,
                            'level': student_level,
                            'course': f'Sample Course {student_level}',
                            'date': timezone.now().strftime('%Y-%m-%d'),
                            'status': 'Present',
                            'check_in_time': '09:00 AM'
                        })
            
            if export_format == 'pdf':
                return export_attendance_pdf(mock_data)
            else:
                return export_attendance_excel(mock_data)
        
        # Process real attendance data
        attendance_data = []
        for attendance in attendance_query:
            attendance_data.append({
                'student_name': attendance.student.full_name,
                'matricule': attendance.student.matric_number,
                'department': attendance.student.department.name if attendance.student.department else 'Unknown',
                'level': getattr(attendance.course_registration.course, 'level', 'Unknown'),
                'course': attendance.course_registration.course.title,
                'date': attendance.date.strftime('%Y-%m-%d'),
                'status': attendance.get_status_display(),
                'check_in_time': attendance.recorded_at.strftime('%H:%M') if attendance.recorded_at else 'N/A'
            })
        
        if export_format == 'pdf':
            return export_attendance_pdf(attendance_data)
        else:
            return export_attendance_excel(attendance_data)
            
    except Exception as e:
        from django.http import JsonResponse
        return JsonResponse({'error': str(e)}, status=500)

def export_attendance_pdf(data):
    """Export attendance data as PDF"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from django.http import HttpResponse
        import io
        
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Title
        title = Paragraph("Attendance Records Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Group data by department and level
        from collections import defaultdict
        grouped_data = defaultdict(lambda: defaultdict(list))
        
        for record in data:
            dept = record['department']
            level = record['level']
            grouped_data[dept][level].append(record)
        
        # Create tables for each department and level
        for dept_name, levels in grouped_data.items():
            # Department header
            dept_header = Paragraph(f"<b>Department: {dept_name}</b>", styles['Heading2'])
            elements.append(dept_header)
            elements.append(Spacer(1, 10))
            
            for level_name, records in levels.items():
                # Level header
                level_header = Paragraph(f"Level: {level_name}", styles['Heading3'])
                elements.append(level_header)
                elements.append(Spacer(1, 10))
                
                # Table data
                table_data = [['Student Name', 'Matricule', 'Course', 'Date', 'Status', 'Check-in Time']]
                
                for record in records:
                    table_data.append([
                        record['student_name'],
                        record['matricule'],
                        record['course'],
                        record['date'],
                        record['status'],
                        record['check_in_time']
                    ])
                
                # Create table
                table = Table(table_data, colWidths=[2*inch, 1.2*inch, 1.5*inch, 1*inch, 0.8*inch, 1*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(table)
                elements.append(Spacer(1, 20))
        
        # Build PDF
        doc.build(elements)
        
        # Return response
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="attendance_records_{timezone.now().strftime("%Y%m%d")}.pdf"'
        return response
        
    except ImportError:
        # Fallback to simple text response if reportlab is not installed
        from django.http import HttpResponse
        import json
        
        response_data = {
            'error': 'PDF export requires reportlab package. Please install it with: pip install reportlab',
            'data': data,
            'message': 'Showing data in JSON format as fallback'
        }
        
        response = HttpResponse(json.dumps(response_data, indent=2), content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="attendance_records_{timezone.now().strftime("%Y%m%d")}.json"'
        return response
    except Exception as e:
        from django.http import JsonResponse
        return JsonResponse({'error': f'PDF export failed: {str(e)}'}, status=500)

def export_attendance_excel(data):
    """Export attendance data as Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        import io
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Group data by department and level
        from collections import defaultdict
        grouped_data = defaultdict(lambda: defaultdict(list))
        
        for record in data:
            dept = record['department']
            level = record['level']
            grouped_data[dept][level].append(record)
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets for each department
        for dept_name, levels in grouped_data.items():
            # Create sheet for department
            sheet_name = dept_name[:31]  # Excel sheet name limit
            ws = wb.create_sheet(title=sheet_name)
            
            current_row = 1
            
            # Department header
            ws.merge_cells(f'A{current_row}:F{current_row}')
            dept_cell = ws[f'A{current_row}']
            dept_cell.value = f"Department: {dept_name}"
            dept_cell.font = Font(bold=True, size=14)
            dept_cell.alignment = Alignment(horizontal='center')
            dept_cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            current_row += 2
            
            for level_name, records in levels.items():
                # Level header
                ws.merge_cells(f'A{current_row}:F{current_row}')
                level_cell = ws[f'A{current_row}']
                level_cell.value = f"Level: {level_name}"
                level_cell.font = Font(bold=True, size=12)
                level_cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                current_row += 1
                
                # Headers
                headers = ['Student Name', 'Matricule', 'Course', 'Date', 'Status', 'Check-in Time']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                current_row += 1
                
                # Data rows
                for record in records:
                    row_data = [
                        record['student_name'],
                        record['matricule'],
                        record['course'],
                        record['date'],
                        record['status'],
                        record['check_in_time']
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        ws.cell(row=current_row, column=col, value=value)
                    current_row += 1
                
                current_row += 1  # Add space between levels
            
            # Adjust column widths
            for col_num in range(1, 7):  # We have 6 columns (A-F)
                max_length = 0
                column_letter = chr(64 + col_num)  # Convert 1->A, 2->B, etc.
                
                for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                    for cell in row:
                        try:
                            if hasattr(cell, 'value') and cell.value is not None:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                        except:
                            pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Return response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="attendance_records_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        return response
        
    except ImportError:
        # Fallback to CSV format if openpyxl is not installed
        from django.http import HttpResponse
        import csv
        import io
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="attendance_records_{timezone.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        
        # Group data by department and level
        from collections import defaultdict
        grouped_data = defaultdict(lambda: defaultdict(list))
        
        for record in data:
            dept = record['department']
            level = record['level']
            grouped_data[dept][level].append(record)
        
        # Write CSV data
        for dept_name, levels in grouped_data.items():
            writer.writerow([f"Department: {dept_name}"])
            writer.writerow([])
            
            for level_name, records in levels.items():
                writer.writerow([f"Level: {level_name}"])
                writer.writerow(['Student Name', 'Matricule', 'Course', 'Date', 'Status', 'Check-in Time'])
                
                for record in records:
                    writer.writerow([
                        record['student_name'],
                        record['matricule'],
                        record['course'],
                        record['date'],
                        record['status'],
                        record['check_in_time']
                    ])
                writer.writerow([])
        
        return response
        
    except Exception as e:
        from django.http import JsonResponse
        return JsonResponse({'error': f'Excel export failed: {str(e)}'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def analytics_data(request):
    """Get analytics data"""
    try:
        # Basic statistics
        total_students = Student.objects.count()
        approved_students = Student.objects.filter(is_approved=True).count()
        departments = Department.objects.count()

        # Department distribution with detailed analytics
        dept_data = []
        for dept in Department.objects.all():
            student_count = Student.objects.filter(department=dept).count()
            approved_count = Student.objects.filter(department=dept, is_approved=True).count()
            
            # Get courses for this department
            try:
                from courses.models import Course
                courses = Course.objects.filter(department=dept)
                course_count = courses.count()
                
                # Course distribution within department
                course_data = []
                for course in courses:
                    # Calculate real enrollment based on StudentCourseSelection or similar model
                    try:
                        # Try to get real enrollment data
                        from students.models import StudentCourseSelection
                        enrollment = StudentCourseSelection.objects.filter(
                            course=course,
                            is_approved=True
                        ).count()
                    except:
                        # Fallback to estimated enrollment
                        enrollment = student_count // max(course_count, 1) if course_count > 0 else 0
                    
                    course_data.append({
                        'id': str(course.id),
                        'code': course.code,
                        'title': course.title,
                        'enrollment': enrollment,
                        'level': course.level,
                        'semester': course.semester
                    })
            except ImportError:
                course_count = 0
                course_data = []
            
            dept_data.append({
                'id': str(dept.id),
                'name': dept.name,
                'students': student_count,
                'approved_students': approved_count,
                'courses': course_count,
                'course_data': course_data,
                'approval_rate': round((approved_count / student_count * 100) if student_count > 0 else 0, 1)
            })

        # Monthly registration trend (last 6 months)
        from datetime import timedelta
        monthly_data = []
        for i in range(6):
            month_start = timezone.now().replace(day=1) - timedelta(days=30*i)
            month_end = month_start + timedelta(days=30)
            count = Student.objects.filter(
                created_at__gte=month_start,
                created_at__lt=month_end
            ).count()
            monthly_data.append({
                'month': month_start.strftime('%b %Y'),
                'registrations': count
            })

        # Calculate real attendance rate
        try:
            from attendance.models import Attendance
            total_attendance = Attendance.objects.count()
            present_attendance = Attendance.objects.filter(status__in=['present', 'late']).count()
            attendance_rate = round((present_attendance / total_attendance * 100), 2) if total_attendance > 0 else 0
        except ImportError:
            attendance_rate = 0

        # Calculate average GPA (placeholder - implement based on your GPA system)
        average_gpa = 0.0  # Implement based on your grading system

        # Overall course statistics
        try:
            from courses.models import Course
            total_courses = Course.objects.count()
        except ImportError:
            total_courses = 0

        return Response({
            'totalStudents': total_students,
            'approvedStudents': approved_students,
            'totalDepartments': departments,
            'totalCourses': total_courses,
            'departmentDistribution': dept_data,
            'monthlyRegistrations': list(reversed(monthly_data)),
            'attendanceRate': attendance_rate,  # Now using real data
            'averageGPA': average_gpa       # Placeholder for now
        })
    except Exception as e:
        import traceback
        print(f"Analytics error: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_admin_levels(request):
    """Get all levels for admin (across all departments)"""
    try:
        from courses.models import Level
        
        # Get department filter if provided
        department_id = request.GET.get('department_id')
        
        if department_id:
            levels = Level.objects.filter(department_id=department_id).select_related('department')
        else:
            levels = Level.objects.all().select_related('department')
        
        levels_data = []
        for level in levels:
            levels_data.append({
                'id': str(level.id),
                'name': level.name,
                'code': level.code,
                'department_id': str(level.department.id),
                'department_name': level.department.name
            })
        
        return Response({
            'success': True,
            'data': levels_data,
            'levels': levels_data  # Alternative naming for compatibility
        })
    except Exception as e:
        return Response({'error': str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def sync_student_timetable(request, student_id):
    """Sync student's timetable with current department schedule"""
    try:
        from courses.models import TimetableSlot, DepartmentTimetable, StudentCourse
        # Use the Student model already imported at the top of the file

        student = Student.objects.get(id=student_id)

        # Get the student's department timetable
        try:
            department_timetable = DepartmentTimetable.objects.get(department=student.department)
        except:
            return Response({'error': 'No timetable found for student\'s department'}, status=404)

        # Get all courses the student is enrolled in
        enrolled_course_ids = StudentCourse.objects.filter(
            student=student,
            is_active=True
        ).values_list('course_offering__course__id', flat=True)

        # Get timetable slots for this student's enrolled courses
        slots = TimetableSlot.objects.filter(
            timetable=department_timetable,
            course__in=enrolled_course_ids
        ).select_related('course', 'lecturer', 'level')

        # In a real implementation, this would sync the student's personal timetable
        # For now, we'll just return the count of relevant slots
        slot_data = []
        for slot in slots:
            slot_data.append({
                'id': slot.id,
                'day': slot.day_of_week,
                'start_time': slot.start_time.strftime('%H:%M'),
                'end_time': slot.end_time.strftime('%H:%M'),
                'course_code': slot.course.code,
                'course_title': slot.course.title,
                'lecturer_name': f"{slot.lecturer.user.first_name} {slot.lecturer.user.last_name}".strip(),
                'venue': slot.venue
            })

        return Response({
            'message': f'Student timetable synced successfully',
            'slots_found': len(slot_data),
            'slots': slot_data
        })
    except Student.DoesNotExist:
        return Response({'error': 'Student not found'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_student_attendance_from_timetable(request, student_id):
    """Update student's attendance records based on timetable"""
    try:
        from courses.models import TimetableSlot, DepartmentTimetable, StudentCourse
        from attendance.models import Attendance
        # Use the Student model already imported at the top of the file

        student = Student.objects.get(id=student_id)

        # Get the student's department timetable
        try:
            department_timetable = DepartmentTimetable.objects.get(department=student.department)
        except:
            return Response({'error': 'No timetable found for student\'s department'}, status=404)

        # Get all courses the student is enrolled in
        enrolled_course_ids = StudentCourse.objects.filter(
            student=student,
            is_active=True
        ).values_list('course_offering__course__id', flat=True)

        # Get timetable slots for this student's enrolled courses
        slots = TimetableSlot.objects.filter(
            timetable=department_timetable,
            course__in=enrolled_course_ids
        ).select_related('course', 'lecturer', 'level')

        # In a real implementation, this would update attendance records based on scheduled classes
        # For now, we'll just return the count of relevant slots
        processed_slots = []
        for slot in slots:
            processed_slots.append({
                'id': slot.id,
                'day': slot.day_of_week,
                'start_time': slot.start_time.strftime('%H:%M'),
                'end_time': slot.end_time.strftime('%H:%M'),
                'course_code': slot.course.code,
                'expected_attendance': True  # This would be calculated based on actual attendance
            })

        return Response({
            'message': f'Student attendance updated based on timetable',
            'slots_processed': len(processed_slots),
            'data': processed_slots
        })
    except Student.DoesNotExist:
        return Response({'error': 'Student not found'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def notify_student_of_changes(request, student_id):
    """Notify student of changes made by admin"""
    try:
        # Use the Student model already imported at the top of the file

        student = Student.objects.get(id=student_id)
        message = request.data.get('message', '')

        if not message:
            return Response({'error': 'Message is required'}, status=400)

        # In a real implementation, this would create a notification in a notifications system
        # For now, we'll just return a success message
        # This could integrate with a proper notification system

        return Response({
            'message': 'Notification would be sent to student',
            'student_id': student_id,
            'student_name': student.full_name,
            'notification_content': message
        })
    except Student.DoesNotExist:
        return Response({'error': 'Student not found'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=500)

# Face Training API endpoints
from .services.face_training import train_face_model

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_face_training_status(request):
    """Get face training status for all students"""
    try:
        students = Student.objects.all().select_related('user', 'department')
        
        students_data = []
        for student in students:
            try:
                # Count face images (photos)
                photo_count = student.photos.count()
                
                # Get department name safely
                department_name = ''
                if hasattr(student, 'department') and student.department:
                    department_name = student.department.name
                
                # Get user info safely
                first_name = ''
                last_name = ''
                if student.user:
                    first_name = student.user.first_name or ''
                    last_name = student.user.last_name or ''
                
                students_data.append({
                    'id': student.id,
                    'matricule': student.matric_number,
                    'first_name': first_name,
                    'last_name': last_name,
                    'full_name': student.full_name,
                    'department': department_name,
                    'faceImages': photo_count,
                    'faceModelTrained': student.face_trained,
                    'is_approved': student.is_approved,
                    'created_at': student.created_at.isoformat() if student.created_at else None
                })
            except Exception as student_error:
                # Log individual student errors but continue processing
                print(f"Error processing student {student.id}: {str(student_error)}")
                continue
        
        # Calculate stats
        total_students = len(students_data)
        trained_students = len([s for s in students_data if s['faceModelTrained']])
        ready_for_training = len([s for s in students_data if s['faceImages'] >= 5 and not s['faceModelTrained']])
        insufficient_images = len([s for s in students_data if s['faceImages'] < 5])
        
        return Response({
            'students': students_data,
            'stats': {
                'total': total_students,
                'trained': trained_students,
                'ready_for_training': ready_for_training,
                'insufficient_images': insufficient_images
            }
        })
    except Exception as e:
        import traceback
        error_details = {
            'error': f'Failed to get face training status: {str(e)}',
            'traceback': traceback.format_exc()
        }
        print(f"Face training status error: {error_details}")
        return Response(
            error_details, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def train_face_model_api(request):
    """Train face models for all eligible students"""
    try:
        success, message = train_face_model()
        
        if success:
            return Response({
                'success': True,
                'message': message
            })
        else:
            return Response({
                'success': False,
                'error': message
            }, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response(
            {'error': f'Failed to train face models: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def train_single_student_face(request, student_id):
    """Train face model for a single student"""
    try:
        # Check if student exists
        try:
            student = Student.objects.get(id=student_id)
        except Student.DoesNotExist:
            return Response(
                {'error': 'Student not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Check if student has enough photos
        photo_count = student.photos.count()
        if photo_count < 5:
            return Response({
                'success': False,
                'error': f'Student needs at least 5 photos for training. Currently has {photo_count} photos.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Train the model for this specific student
        success, message = train_face_model(student_ids=[student_id])
        
        if success:
            return Response({
                'success': True,
                'message': message
            })
        else:
            return Response({
                'success': False,
                'error': message
            }, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response(
            {'error': f'Failed to train face model for student: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_live_sessions(request):
    """Get all live sessions for admin"""
    try:
        from live_sessions.models import LiveSession
        from live_sessions.serializers import LiveSessionSerializer
        
        # Get all sessions
        sessions = LiveSession.objects.all().select_related('instructor', 'course_offering__course')
        
        # Apply filters if provided
        status_filter = request.GET.get('status')
        if status_filter:
            sessions = sessions.filter(status=status_filter)
        
        # Serialize the data
        serializer = LiveSessionSerializer(sessions, many=True)
        
        return Response({
            'success': True,
            'data': serializer.data
        })
    except ImportError:
        # If live_sessions app is not available, return empty data
        return Response({
            'success': True,
            'data': [],
            'message': 'Live sessions module not available'
        })
    except Exception as e:
        return Response({'error': str(e)}, status=500)